import tkinter
from pprint import pprint
from tkinter import filedialog
import os
import pandas as pd
from openpyxl.utils import get_column_letter

root = tkinter.Tk()
root.withdraw()  # use to hide tkinter window


def prepare_sheet(sheet):
    merged_cells = list(map(str, sheet.merged_cells.ranges))  # Получаю список объединенных диапазонов
    # Разъединяю объединенные ячейки и дублирую запись
    for item in merged_cells:
        sheet.unmerge_cells(item)
        merged_cells_range = item.split(":")
        if merged_cells_range[0][0] == merged_cells_range[1][0]:
            letter = item.split(":").pop(0)[0]  # Символ столбца диапазона
            start = int(item.split(":").pop(0)[1:])  # Начало диапазона
            end = int(item.split(":").pop()[1:])  # Конец диапазона
            copy_cell = sheet[(letter + str(start))].value
            for n in range(start, end + 1):
                cell = letter + str(n)
                sheet[cell].value = copy_cell


def restore_rnpd_file_format(sheet):
    for row_index, row in enumerate(sheet):
        for cell_index, cell in enumerate(row):
            if (row_index > 0):
                row[6].value = str(row[6].value).replace(',', '.')
                if (len(str(row[21].value)) > 1):
                    row[21].value = str(row[21].value).rjust(10, "0")
                row[22].value = str(row[22].value).rjust(6, "0")
                row[24].value = str(row[24].value).rjust(6, "0")
                row[25].value = str(row[25].value).rjust(3, "0")


def search_for_file_path(file_name):
    currdir = os.getcwd()
    temp_dir = filedialog.askdirectory(parent=root, initialdir=currdir,
                                       title=f'Please select a directory of file: {file_name}')
    if len(temp_dir) > 0:
        print("You chose: %s" % temp_dir)
    return temp_dir

'''
def get_headers_row_index(worksheet):
    not_empty_cells_count = 0
    headers_row_index = 0
    for row_index, row in enumerate(worksheet):
        for cell_index, cell in enumerate(row):
            if (cell.value != None):
                not_empty_cells_count = not_empty_cells_count + 1

        if (not_empty_cells_count >= 2):
            headers_row_index = row_index
            break
    #print(f"Worksheet {worksheet} headers row index: {headers_row_index}")
    return headers_row_index
'''
def get_headers_row_index(worksheet):
    headers_row_index = 0
    for row_index, row in enumerate(worksheet, start=1):
        row_not_empty = not row_is_empty(worksheet, row_index)
        #print(f'Row: {row_index} not empty: {row_not_empty}')
        if row_not_empty:
            headers_row_index = row_index
            break
    print(f"Worksheet {worksheet} headers row index: {headers_row_index}")
    return headers_row_index


def row_is_empty(worksheet, row_index):
    not_empty_cells_count = 0

    for cell in worksheet[row_index][1:5]:
        #print(cell.value)
        if (cell.value != None):
            not_empty_cells_count = not_empty_cells_count + 1

    #print(f'For row {row_index} not_empty_cells_count: {not_empty_cells_count}')
    if (not_empty_cells_count > 2):
        return False
    else:
        return True


def get_last_row_index(worksheet, headers_row_index = -1): #работает медленно
    last_row_index = 0

    if(headers_row_index == -1):
        headers_row_index = get_headers_row_index(worksheet)

    for row_index, row in enumerate(worksheet, start=headers_row_index):
        if (row_is_empty(worksheet, row_index)):
            break

        last_row_index = row_index

    return last_row_index

def get_last_row_index_alternative(worksheet, column_index,  headers_row_index = -1): #работает быстрее get_last_row_index()
    last_row_index = 0

    if(headers_row_index == -1):
        headers_row_index = get_headers_row_index(worksheet)

    column_values = []
    for row_index, row in enumerate(worksheet, start=headers_row_index):
        if (row_index > headers_row_index):
            column_values.append(str(row[column_index].value))
    return headers_row_index + len(column_values)

def get_last_column_index(worksheet,headers_row_index = -1 , empty_columns_count =3):
    if(headers_row_index == -1):
        headers_row_index = get_headers_row_index(worksheet)

    first_column = worksheet[headers_row_index][1].col_idx
    #print(f'First column index: {first_column}')
    column_count = 0
    empty_columns_founded = 0

    for cell in worksheet[headers_row_index]:
        if cell.value != None:
            column_count = column_count + 1
        else:
            empty_columns_founded = empty_columns_founded + 1
            if empty_columns_founded == empty_columns_count:
                break

    last_column_index = first_column + column_count - (empty_columns_count -1)
    print(f'{worksheet} last column index: {last_column_index}')
    return last_column_index

def get_headers_names(worksheet, headers_row_index = -1):
    column_names = []

    if(headers_row_index <= -1):
        headers_row_index = get_headers_row_index(worksheet)

    for cell in worksheet[headers_row_index]:
        column_names.append(cell.value)

    column_names = [x.lower() for x in column_names if x is not None]
    #pprint(column_names)
    return column_names

def get_index_any_from_list(list, value_list):
    value_index = None
    for v in value_list:
        v_lower = v.lower()
        if(v_lower in map(str.lower, list)):
            value_index = list.index(v_lower)
            break
        #else:
        #    print(f'Didn`t find any from: {value_list} in {list}')
    return value_index

def convert_XLSB_to_XLSX_and_get_new_filepath_through_pandas(xlsb_file_path, sheet_name):
    rnpd_file = pd.read_excel(xlsb_file_path, engine='pyxlsb', sheet_name=sheet_name)
    xlsx_file_path = xlsb_file_path.split(".")[0] + "_result" + ".xlsx"
    rnpd_file.to_excel(xlsx_file_path, index=False)
    return xlsx_file_path


def get_cell_range(worksheet, min_row, max_row, min_column, max_column):
    cell_range = [worksheet.cell(row=i, column=j).value for i in range(min_row, max_row) for j in
                  range(min_column, max_column)]
    return cell_range


def get_worksheet_cell_range(worksheet):
    cell_range = [worksheet.cell(row=i, column=j).value for i in range(worksheet.min_row, worksheet.max_row) for j in
                  range(worksheet.min_column, worksheet.max_column)]
    return cell_range


def binary_search(arr, x):  # https://www.freecodecamp.org/news/how-to-search-large-datasets-in-python/
    '''
    Бинарный поиск - работает только с отсортированными списками. Быстрее чем перебор, но при этом сортировка - тяжелая операция, поэтому нужно смотреть будет ли быстрее перебора.
    '''
    low = 0
    high = len(arr) - 1
    while low <= high:
        mid = (low + high) // 2
        if arr[mid] < x:
            low = mid + 1
        elif arr[mid] > x:
            high = mid - 1
        else:
            return mid
    return -1