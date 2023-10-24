#!/usr/bin/python
# -*- coding: utf8 -*-

import os
from pprint import pprint
from tkinter import filedialog

import openpyxl
from openpyxl.styles import PatternFill

import MyClasses
import MyFunctions

DEFAULT_COLOR = "00FFFF00"
WARNING_COLOR = '00FF0000'
DEFAULT_FILL_TYPE = "solid"

# Указываем пути к исходным файлам
rnpd_file_path = filedialog.askopenfilename(title="Выберите файл РНПД")
cfo140_file_path = filedialog.askopenfilename(title="Выберите файл 014ЦФО")

cfo140_new_file_path = os.path.splitext(cfo140_file_path)[0] + "_result" + ".xlsx"

# Проверяем что пути к файлам указаны
if (len(rnpd_file_path) == 0 or len(cfo140_file_path) == 0):
    print("Файлы не выбраны. Перезапустите программу и выберете файлы!")
    raise SystemExit

# Загружаем данные из файла cfo140
cfo140_file = openpyxl.open(cfo140_file_path, data_only=False, read_only=False)
working_sheet_cfo140_file = cfo140_file.active
MyFunctions.prepare_sheet(working_sheet_cfo140_file)  # Убираем объединенные ячейки в файле 014ЦФО

# Начинаем обрабатывать файлы
rnpd_new_file_path = MyFunctions.convert_XLSB_to_XLSX_and_get_new_filepath_through_pandas(rnpd_file_path,
                                                                                          'Шаблон')  # я так понял что могу только считывать данные из файла xlxb, а менять ничего не могу там, поэтому делаю копию в другом формате
rnpd_file_result = openpyxl.open(rnpd_new_file_path, data_only=False,
                                 read_only=False)  # сюда буду записывать все изменения файла РНПД, отсюда сотрудник потом сможет перекопировать данные в файл .xlxb
rnpd_file_result_ws = rnpd_file_result.active
rnpd_file_result_ws.title = 'Шаблон'
MyFunctions.restore_rnpd_file_format(rnpd_file_result_ws)

matched_item_list = []  # сюда складываю все совпадения по индексу

cfo140_headers_row_index = MyFunctions.get_headers_row_index(working_sheet_cfo140_file)
rnpd_file_headers_row_index = MyFunctions.get_headers_row_index(rnpd_file_result_ws)

cfo140_last_row = MyFunctions.get_last_row_index(working_sheet_cfo140_file, cfo140_headers_row_index)
rnpd_last_row = MyFunctions.get_last_row_index(rnpd_file_result_ws, rnpd_file_headers_row_index)

cfo140_last_column_index = MyFunctions.get_last_column_index(working_sheet_cfo140_file, cfo140_headers_row_index)
result_colunm_index = cfo140_last_column_index + 1
working_sheet_cfo140_file.insert_cols(idx=result_colunm_index)
working_sheet_cfo140_file.cell(cfo140_headers_row_index, result_colunm_index).value = 'Результат обработки'
working_sheet_cfo140_file.cell(cfo140_headers_row_index, result_colunm_index).fill = PatternFill(DEFAULT_FILL_TYPE,
                                                                                                 start_color=DEFAULT_COLOR)
cfo140_file.save(cfo140_new_file_path)
cfo140_file = openpyxl.open(cfo140_new_file_path, data_only=False, read_only=False)
working_sheet_cfo140_file = cfo140_file.active


cfo140_column_names = MyFunctions.get_headers_names(working_sheet_cfo140_file, cfo140_headers_row_index)
rnpd_file_column_names = MyFunctions.get_headers_names(rnpd_file_result_ws, rnpd_file_headers_row_index)

cfo140_column_index_nomer_dogovora = MyFunctions.get_index_any_from_list(cfo140_column_names, (
'Системный номер договора', 'Системный № Договора', ' Системный № Договора'))
cfo140_column_index_statja_zatrat = MyFunctions.get_index_any_from_list(cfo140_column_names, ('Статья затрат', 'СЗ'))
cfo140_column_index_SHPP = MyFunctions.get_index_any_from_list(cfo140_column_names, ('ШПП', 'ШПЗ'))
cfo140_column_index_BP = MyFunctions.get_index_any_from_list(cfo140_column_names, ('БП', 'Бизнес процесс'))
cfo140_column_index_BE = MyFunctions.get_index_any_from_list(cfo140_column_names, ('БЕ', 'Код ОЕ', 'ОЕ'))
cfo140_column_index_CFO = MyFunctions.get_index_any_from_list(cfo140_column_names, ('ЦФО-получатель', 'ЦФО'))
cfo140_column_index_Kontragent = MyFunctions.get_index_any_from_list(cfo140_column_names,
                                                                     ('Наименование контрагента', 'Поставщик'))
cfo140_column_index_Summa = MyFunctions.get_index_any_from_list(cfo140_column_names, (
'Сумма без НДС (руб)', 'Плановая сумма поставки (без НДС)', 'Плановая сумма поставки (без НДС)'))
cfo140_column_index_Data = MyFunctions.get_index_any_from_list(cfo140_column_names, (
'Дата фактического оказания услуг', 'Дата факт. оказ. Услуги'))
cfo140_column_index_Subject = cfo140_column_names.index('Субъект'.lower())
cfo140_column_index_Programma = cfo140_column_names.index('Программа'.lower())
# cfo140_column_index_Programma = MyFunctions.get_index_any_from_list(cfo140_column_names, ('Программа','программа'))
cfo140_column_index_MVZ = cfo140_column_names.index('МВЗ'.lower())
# cfo140_column_index_Kommentariy = MyFunctions.get_index_any_from_list(cfo140_column_names, ('Комментарий','комментарий'))
cfo140_column_index_Kommentariy = cfo140_column_names.index('Результат обработки'.lower())

rnpd_column_index_nomer_dogovora = rnpd_file_column_names.index('№ Договора'.lower())
rnpd_column_index_statja_zatrat = rnpd_file_column_names.index('СЗ'.lower())
rnpd_column_index_SHPZ = rnpd_file_column_names.index('ШПЗ'.lower())
rnpd_column_index_subject = rnpd_file_column_names.index('Субъект'.lower())
rnpd_column_index_BP = rnpd_file_column_names.index('Бизнес процесс'.lower())
rnpd_column_index_summa_raspredelenija = rnpd_file_column_names.index('Сумма строки распределения'.lower())
rnpd_column_index_OE = rnpd_file_column_names.index('Код ОЕ'.lower())

rnpd_OE_values = []
for rpnd_row_index, rnpd_row in enumerate(rnpd_file_result_ws, start=rnpd_file_headers_row_index):
    if (rpnd_row_index > rnpd_file_headers_row_index):
        rnpd_OE_values.append(str(rnpd_row[rnpd_column_index_OE].value).rjust(5, "0"))

for cfo140_row_index, cfo140_row in enumerate(working_sheet_cfo140_file, start=1):
    if (cfo140_row[cfo140_column_index_BE].value not in rnpd_OE_values):
        continue
    # if (cfo140_row_index > cfo140_headers_row_index and not MyFunctions.row_is_empty(working_sheet_cfo140_file, cfo140_row_index)):  # пропускаю первую пустую строку и заголовок в файле
    if (cfo140_row_index > cfo140_headers_row_index and cfo140_row_index <= cfo140_last_row):  # пропускаю заголовок в файле
        cfo140_row_key = str(cfo140_row[cfo140_column_index_nomer_dogovora].value) + '-' + str(
            cfo140_row[cfo140_column_index_statja_zatrat].value) + '-' + str(
            cfo140_row[cfo140_column_index_SHPP].value) + '-' + str(
            cfo140_row[cfo140_column_index_Subject].value) + '-' + str(
            cfo140_row[cfo140_column_index_BP].value)

        for rpnd_row_index, rnpd_row in enumerate(rnpd_file_result_ws, start=1):
            # print(f'ЦФО индекс: {cfo140_row_index}, РНПД индекс: {rpnd_row_index}')
            for cell_index, cell in enumerate(rnpd_row):
                if (rpnd_row_index > rnpd_file_headers_row_index and rpnd_row_index <= rnpd_last_row):
                    rnpd_row_key = str(
                        rnpd_row[rnpd_column_index_nomer_dogovora].value.split()[0]) + '-' + str(
                        rnpd_row[rnpd_column_index_statja_zatrat].value) + '-' + str(
                        rnpd_row[rnpd_column_index_SHPZ].value) + '-' + str(
                        rnpd_row[rnpd_column_index_subject].value).rjust(6, "0") + '-' + str(
                        rnpd_row[rnpd_column_index_BP].value).rjust(6, "0")
                    #print(f'rnpd_row_key: {rnpd_row_key}, cfo140_row_key: {cfo140_row_key},')
                    if (cfo140_row_key == rnpd_row_key and cell_index == rnpd_column_index_summa_raspredelenija):
                        old_summ = rnpd_row[rnpd_column_index_summa_raspredelenija].value
                        new_summ = cfo140_row[cfo140_column_index_Summa].value
                        matched_item = MyClasses.MatchItem(cfo140_row_index, rpnd_row_index, old_summ, new_summ,
                                                           cfo140_row_key)
                        matched_item_list.append(matched_item)
                        # Обрабатываю сопадения после цикла


for cfo140_row_index, cfo140_row in enumerate(working_sheet_cfo140_file, start=1):
    print(f'Row index: {cfo140_row_index}, header row index: {cfo140_headers_row_index}, last row: {cfo140_last_row}')
    if (cfo140_row_index > cfo140_headers_row_index and cfo140_row_index <= cfo140_last_row):  # пропускаю первую пустую строку и заголовок в файле
        if (cfo140_row[cfo140_column_index_BE].value not in rnpd_OE_values):
            cfo140_row[cfo140_column_index_Kommentariy].value = f"Файл РНПД не содержит БЕ: {cfo140_row[cfo140_column_index_BE].value}!"
            cfo140_row[cfo140_column_index_Kommentariy].fill = PatternFill(DEFAULT_FILL_TYPE, start_color=WARNING_COLOR)
            continue

        cfo140_row_key = str(cfo140_row[cfo140_column_index_nomer_dogovora].value) + '-' + str(
            cfo140_row[cfo140_column_index_statja_zatrat].value) + '-' + str(
            cfo140_row[cfo140_column_index_SHPP].value) + '-' + str(
            cfo140_row[cfo140_column_index_Subject].value) + '-' + str(
            cfo140_row[cfo140_column_index_BP].value)


        if(any(x.key == cfo140_row_key for x in matched_item_list)):
            row_key_matches_count = list(map(lambda x: x.key, matched_item_list)).count(cfo140_row_key)
            if (row_key_matches_count > 1):
                rnpd_rows = list(filter(lambda x: cfo140_row_key in x.key, matched_item_list))
                rows_numbers = list(map(lambda x: x.rpnd_row_index, rnpd_rows))  # номера строк в файле РНПД

                cfo140_row[cfo140_column_index_Kommentariy].value = f"Найдено более одного сопадения в файле РНПД. Строки: {', '.join(map(str, rows_numbers))}!"
                cfo140_row[cfo140_column_index_Kommentariy].fill = PatternFill(DEFAULT_FILL_TYPE, start_color=DEFAULT_COLOR)

                '''
                #2)	Найдено более одного совпадения. Необходимо только записать Результат. В результате указать, что найдено более одного совпадения, указать номера строк из шаблона РНПД, строки подкрасить светло красноватым не ярким цветом.
                '''
                cfo140_file.save(cfo140_new_file_path)
            else:
                matched_item = list(filter(lambda x: x.key == cfo140_row_key, matched_item_list))[0]

                if (matched_item.old_summ != matched_item.new_summ):
                    rnpd_cell = rnpd_file_result_ws.cell(matched_item.rpnd_row_index + 1,
                                                         rnpd_column_index_summa_raspredelenija + 1)
                    rnpd_cell.value = matched_item.new_summ
                    rnpd_cell.fill = PatternFill(DEFAULT_FILL_TYPE, start_color=DEFAULT_COLOR)

                    cfo140_row[
                        cfo140_column_index_Kommentariy].value = f"В файле РНПД отредактирована сумма по строке: {matched_item.rpnd_row_index + 1}. Старая сумма: {matched_item.old_summ}. Новая сумма {matched_item.new_summ}."
                else:
                    cfo140_row[
                        cfo140_column_index_Kommentariy].value = f"В файле РНПД сумма по строке: {matched_item.rpnd_row_index + 1} совпадает."

                cfo140_file.save(cfo140_new_file_path)
                '''
                #1)	Найдено одно совпадение. Необходимо в шаблоне РНПД на листе Шаблон поменять только сумму в колонке R, сумма строки распределения. Сумму указываем из реестра по 014 ЦФО колонка F, сумма без НДС. Результат записать в реестр по 014 ЦФО, колонку можно создать справа таблицы. В результате указать, что найдено совпадение, указать 
                '''
        else:
            BE = cfo140_row[cfo140_column_index_BE].value
            CFO = cfo140_row[cfo140_column_index_CFO].value
            KONTRAGENT = cfo140_row[cfo140_column_index_Kontragent].value
            DOGOVOR = cfo140_row[cfo140_column_index_nomer_dogovora].value
            SUMMA = cfo140_row[cfo140_column_index_Summa].value
            # print(f'Row: {cfo140_row_index} - Date: {cfo140_row[cfo140_column_index_Data].value}')
            DATA = cfo140_row[cfo140_column_index_Data].value.strftime("%d.%m.%Y")
            SHPP = cfo140_row[cfo140_column_index_SHPP].value
            SZ = cfo140_row[cfo140_column_index_statja_zatrat].value
            PROGRAMMA = str(cfo140_row[cfo140_column_index_Programma].value).rjust(10, "0")
            SUBJECT = str(cfo140_row[cfo140_column_index_Subject].value).rjust(6, "0")
            BP = str(cfo140_row[cfo140_column_index_BP].value).rjust(6, "0")
            MVZ = cfo140_row[cfo140_column_index_MVZ].value

            row = [BE, CFO, '', KONTRAGENT, '', DOGOVOR, '', '', '', '', '', '', '', DATA, '', '', '', SUMMA, SHPP,
                   SZ,
                   '', PROGRAMMA, SUBJECT, '', BP, MVZ, '', '', '', '', '', '', '', '', '', '', '', '', '', '']

            rnpd_file_result_ws.append(row)
            rnpd_file_result.save(rnpd_new_file_path)

            for cell_index, cell in enumerate(rnpd_file_result_ws[rnpd_file_result_ws.max_row], start=1):
                rnpd_cell = rnpd_file_result_ws.cell(rnpd_file_result_ws.max_row, cell_index)
                rnpd_cell.fill = PatternFill(DEFAULT_FILL_TYPE, start_color=DEFAULT_COLOR)

            cfo140_row[
                cfo140_column_index_Kommentariy].value = f"В файле РНПД не найдено совпадений. Добавлена строка: {rnpd_file_result_ws.max_row}!"
            cfo140_row[cfo140_column_index_Kommentariy].fill = PatternFill(DEFAULT_FILL_TYPE, start_color=DEFAULT_COLOR)

            '''
            #Не найдено ни одного совпадения. В шаблоне РНПД необходимо завести новую строку, перенести данные из реестра 014 ЦФО.  Соответствия столбцов указаны в таблице ниже.
            #В результате указать, что не найдено совпадение, создана строка, указать номер строки из шаблона РНПД,  строку подкрасить желтым
            '''

        cfo140_file.save(cfo140_new_file_path)

rnpd_file_result.save(rnpd_new_file_path)
cfo140_file.save(cfo140_new_file_path)

# Запускаю сформированные файлы
os.startfile(rnpd_new_file_path)
os.startfile(cfo140_new_file_path)
